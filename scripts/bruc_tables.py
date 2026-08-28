"""Build the Nigeria brucellosis (WAHIS) table set: one XLSX workbook + per-table CSVs
+ a Markdown version of every table. Every figure is computed from the raw CSV."""
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from bruc_data import (RECORDS, YEAR_START, YEAR_END, TOTAL_STATES, TOTAL_OUTBREAKS,
                       RECORD_COUNT, YEARS_SPAN, YEARS_REPORTING, STATES_REPORTING,
                       PERIODS_REPORTED, PERIODS_POSSIBLE, DUPLICATE_RECORDS,
                       ANIMAL_CATEGORY, COUNTRY, SOURCE_FILENAME,
                       FIELD_COMPLETENESS, ZONE_STATES,
                       by_year, by_state, by_zone, by_subtype, by_semester)

# Repository root: one level above scripts/
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Generated table outputs
BASE = PROJECT_ROOT
CSV_DIR = os.path.join(BASE, "tables")
os.makedirs(CSV_DIR, exist_ok=True)

SEM_LABEL = {"H1": "H1 (Jan–Jun)", "H2": "H2 (Jul–Dec)"}
SRC = (f"World Organisation for Animal Health (WOAH), WAHIS quantitative data — {COUNTRY}, "
       f"Brucellosis. Extract: “{SOURCE_FILENAME}”. {RECORD_COUNT} records, "
       f"{TOTAL_OUTBREAKS} reported new outbreaks, {YEAR_START}–{YEAR_END}.")
PCT = f"% of {TOTAL_OUTBREAKS} outbreaks"

d_year, n_year = by_year()
d_state, n_state = by_state()
d_zone, rep_zone = by_zone()
d_sub, n_sub = by_subtype()
d_sem, n_sem = by_semester()

zone_of = {s: z for z, ss in ZONE_STATES.items() for s in ss}


def pct(v):
    return round(v / TOTAL_OUTBREAKS * 100, 1)


# ---------------------------------------------------------------- table specs
t_summary = (
    ["Metric", "Value", "What it is not"],
    [
        ["WAHIS quantitative records", RECORD_COUNT,
         "Not the outbreak count — one record can carry a New outbreaks value > 1"],
        ["Reported new outbreaks (sum of New outbreaks)", TOTAL_OUTBREAKS,
         f"Not {TOTAL_OUTBREAKS} cases and not {TOTAL_OUTBREAKS} infected animals — WAHIS "
         "reports no case- or animal-level counts in this extract"],
        ["States represented", f"{STATES_REPORTING} of {TOTAL_STATES} (36 states + FCT)",
         "Reporting coverage, not disease prevalence"],
        ["Years with at least one record",
         f"{YEARS_REPORTING} of {YEARS_SPAN} ({YEAR_START}–{YEAR_END})",
         "A year with no record means no WAHIS record was identified, not that zero disease occurred"],
        ["Half-year reporting periods represented", f"{PERIODS_REPORTED} of {PERIODS_POSSIBLE}",
         f"Reporting coverage, not a {PERIODS_REPORTED}-event count"],
        ["Brucella species / categories present", len(d_sub),
         "Derived from the Disease field, not from the (empty) Serotype/Subtype/Genotype field"],
        ["Animal category", ANIMAL_CATEGORY, "Constant across every record in the extract"],
        ["Duplicate records (state × year × semester)", DUPLICATE_RECORDS, ""],
        ["Jan–Jun (H1) outbreaks", d_sem["H1"],
         "A distribution by reporting semester, not a seasonality finding"],
        ["Jul–Dec (H2) outbreaks", d_sem["H2"], "Same caveat"],
    ],
    "01-summary-metrics",
    "Summary metrics",
)

t_annual = (
    ["Year", "Reported new outbreaks", "Records", "Note"],
    [[y, d_year[y] if d_year[y] else "—", n_year[y] if n_year[y] else "—",
      "" if d_year[y] else "No WAHIS record identified for this year"]
     for y in range(YEAR_START, YEAR_END + 1)]
    + [["Total", TOTAL_OUTBREAKS, RECORD_COUNT, ""]],
    "02-annual-outbreaks",
    "Annual outbreak totals",
)

t_states = (
    ["Rank", "State", "Geopolitical zone (derived)", "Reported new outbreaks", "Records", PCT],
    [[i + 1, s, zone_of[s], v, n_state[s], pct(v)] for i, (s, v) in enumerate(d_state.items())]
    + [["", f"Total ({STATES_REPORTING} states)", "", TOTAL_OUTBREAKS, RECORD_COUNT, 100.0]],
    "03-state-outbreaks",
    "State-level outbreak totals",
)

t_zones = (
    ["Geopolitical zone (derived)", "Reported new outbreaks", PCT,
     "States reporting", "States in zone", "Reporting states"],
    [[z, v, pct(v), len(rep_zone[z]), len(ZONE_STATES[z]),
      ", ".join(sorted(rep_zone[z])) or "none"] for z, v in d_zone.items()]
    + [["Total", TOTAL_OUTBREAKS, 100.0, STATES_REPORTING, TOTAL_STATES, ""]],
    "04-zone-outbreaks-derived",
    "Zone totals (DERIVED — not WAHIS-native)",
)

t_sub = (
    ["Brucella species / category (from the Disease field)", "Reported new outbreaks",
     "Records", PCT],
    [[f"{s} (Inf. with)", v, n_sub[s], pct(v)] for s, v in d_sub.items()]
    + [["Total", TOTAL_OUTBREAKS, RECORD_COUNT, 100.0]],
    "05-brucella-species",
    "Brucella species / category",
)

t_sem = (
    ["Reporting semester", "Reported new outbreaks", "Records", PCT],
    [[SEM_LABEL[s], d_sem[s], n_sem[s], pct(d_sem[s])] for s in ("H1", "H2")]
    + [["Total", TOTAL_OUTBREAKS, RECORD_COUNT, 100.0]],
    "06-semester-split",
    "Distribution by reporting semester",
)


def field_note(field, populated):
    if field == "Disease":
        return "Varies per record; source of the species/category breakdown"
    if field == "Animal Category":
        return f"Constant: “{ANIMAL_CATEGORY}”"
    if not populated:
        note = "Unpopulated in every record"
        if field.startswith("Serotype"):
            note += (" — NOT the source of the species breakdown; that comes from "
                     "the Disease field")
        return note
    return ""


t_fields = (
    ["WAHIS field", "Completeness", f"Populated records (of {RECORD_COUNT})", "Note"],
    [[f, f"{p}%", round(RECORD_COUNT * p / 100), field_note(f, p)]
     for f, p in FIELD_COMPLETENESS],
    "07-field-completeness",
    "WAHIS field completeness",
)

t_records = (
    ["#", "Year", "Semester", "Period", "State", "Geopolitical zone (derived)",
     "Disease (Brucella species/category)", "Animal category", "New outbreaks"],
    [[i + 1, y, s, f"{'Jan–Jun' if s == 'H1' else 'Jul–Dec'} {y}", st, zone_of[st],
      f"{sub} (Inf. with)", ANIMAL_CATEGORY, o]
     for i, (y, s, st, sub, o) in enumerate(RECORDS)]
    + [["", "", "", "", "", "", "", "Total", TOTAL_OUTBREAKS]],
    "08-wahis-records",
    f"The {RECORD_COUNT} WAHIS records",
)

years = list(range(YEAR_START, YEAR_END + 1))
matrix = {(st, y): 0 for st in d_state for y in years}
for y, _s, st, _sub, o in RECORDS:
    matrix[(st, y)] += o
t_matrix = (
    ["State"] + [str(y) for y in years] + ["Total"],
    [[st] + [matrix[(st, y)] or "" for y in years] + [d_state[st]] for st in d_state]
    + [["Total"] + [d_year[y] or "" for y in years] + [TOTAL_OUTBREAKS]],
    "09-state-by-year-matrix",
    "Reporting continuity: state × year (blank = no WAHIS record identified)",
)

TABLES = [t_summary, t_annual, t_states, t_zones, t_sub, t_sem, t_fields, t_records, t_matrix]

# ------------------------------------------------------------------ CSV + MD
md = ["# Nigeria brucellosis — WAHIS tables", "", SRC, "",
      "Zone groupings are a standard Nigerian administrative classification applied to roll up",
      "WAHIS state-level records; WAHIS itself reports state (Administrative Division) only.", ""]
for header, rows, slug, title in TABLES:
    with open(os.path.join(CSV_DIR, f"{slug}.csv"), "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)
    md += [f"## {title}", "",
           "| " + " | ".join(str(h) for h in header) + " |",
           "|" + "|".join("---" for _ in header) + "|"]
    md += ["| " + " | ".join(str(c) for c in r) + " |" for r in rows]
    md += [""]
with open(os.path.join(CSV_DIR, "brucellosis-tables.md"), "w", encoding="utf-8") as fh:
    fh.write("\n".join(md))

# --------------------------------------------------------------------- XLSX
HEAD_FILL = PatternFill("solid", fgColor="2A78D6")
TOTAL_FILL = PatternFill("solid", fgColor="EDF3FC")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10.5)
THIN = Side(style="thin", color="DDDCD8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()
ws0 = wb.active
ws0.title = "Read me"
readme = [
    (f"{COUNTRY} brucellosis surveillance — WAHIS tables", True),
    ("", False),
    ("Source", True),
    (SRC, False),
    ("Every figure in this workbook is computed directly from that CSV.", False),
    ("", False),
    ("Five concepts kept deliberately distinct", True),
    (f"1. Record count ({RECORD_COUNT}) — rows in the WAHIS quantitative extract; one row "
     "per state × half-year.", False),
    (f"2. Outbreak count ({TOTAL_OUTBREAKS}) — the sum of the New outbreaks field. "
     "NOT cases, animals, or herds.", False),
    (f"3. Reporting coverage — {STATES_REPORTING}/{TOTAL_STATES} states, "
     f"{YEARS_REPORTING}/{YEARS_SPAN} years, {PERIODS_REPORTED}/{PERIODS_POSSIBLE} half-year "
     "periods. A coverage measure.", False),
    (f"4. Field completeness — the share of the {RECORD_COUNT} records in which a field is "
     "populated. A data-quality measure.", False),
    ("5. True disease burden — incidence, prevalence, attack rate, mortality, case fatality, "
     "vaccination", False),
    ("   coverage. NONE of these can be computed from this extract; the denominators are absent.", False),
    ("", False),
    ("Cautions", True),
    ("• Absence of a WAHIS record is not absence of disease. It may be a reporting, "
     "surveillance,", False),
    ("  diagnostic or submission gap — this extract cannot distinguish between them.", False),
    ("• Geographic concentration (Plateau, Kaduna, North Central, North West) describes "
     "reported", False),
    ("  outbreak distribution, not confirmed epidemiological hotspots.", False),
    (f"• The {d_sem['H1']} / {d_sem['H2']} Jan–Jun vs Jul–Dec split is a "
     "distribution by reporting semester, not seasonality.", False),
    ("• Geopolitical-zone totals are DERIVED. WAHIS reports Administrative Division "
     "(state) only.", False),
    (f"• The dedicated Serotype/Subtype/Genotype field is empty in all {RECORD_COUNT} "
     "records. The species", False),
    ("  breakdown comes from the Disease field, which is a different field and does vary "
     "per record.", False),
    ("• This analysis is independent and does not represent an official position of WOAH "
     "or the", False),
    ("  Nigerian government.", False),
    ("", False),
    ("Sheets", True),
] + [(f"• {title}", False) for _h, _r, _s, title in TABLES]
for i, (text, bold) in enumerate(readme, start=1):
    c = ws0.cell(row=i, column=1, value=text)
    c.font = Font(bold=bold, size=13 if i == 1 else 10.5)
    c.alignment = Alignment(vertical="center")
ws0.column_dimensions["A"].width = 112
ws0.sheet_view.showGridLines = False

SHEET_NAMES = ["Summary", "Annual", "By state", "By zone (derived)", "Species",
               "Semester", "Field completeness", "Records", "State x year"]
for (header, rows, _slug, title), name in zip(TABLES, SHEET_NAMES):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value=SRC).font = Font(size=8.5, color="8A8880")
    for j, h in enumerate(header, start=1):
        c = ws.cell(row=4, column=j, value=h)
        c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BORDER
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for i, row in enumerate(rows, start=5):
        is_total = any(str(c).strip().lower().startswith("total") for c in row)
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(size=10.5, bold=is_total)
            c.border = BORDER
            if is_total:
                c.fill = TOTAL_FILL
            if j > 1 and isinstance(v, (int, float)):
                c.alignment = Alignment(horizontal="center")
    for j in range(1, len(header) + 1):
        longest = max([len(str(header[j - 1]))] + [len(str(r[j - 1])) for r in rows])
        ws.column_dimensions[get_column_letter(j)].width = min(max(longest + 3, 9), 62)
    ws.row_dimensions[4].height = 32
    ws.freeze_panes = "A5"

wb.save(os.path.join(BASE, "brucellosis-tables.xlsx"))
print("tables written to", BASE)
for f in sorted(os.listdir(CSV_DIR)):
    print("  tables/" + f)
